import openpyxl # Это модуль Python для чтения и записи Excel 2010 xlsx/xlsm/xltx/xltm файлов
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import os
import datetime
import time
import subprocess, sys

try:
    book = openpyxl.Workbook() # Создание пустого документа
    # book = openpyxl.load_workbook('sample.xlsx') # Открытие существующего документа
    sheet = book.active # Активация документа
    # sheet['A1'] = 42 # Выдача текста в ячейку
    # sheet.append([1, 2, 3, 4]) # Добавление строки с несколькими значениями
    # sheet['B3'] = datetime.datetime.now() # Вставка даты в ячейку
    # sheet['A3'] = time.strftime("%x") # Вставка времени в ячейку
    # sheet.cell(row=4, column=5).value = 77 # Вставка значения в ячейку
    # Кортедж
    rows = [
        ['1.jpeg', 'MERCEDES BENZ CLA CLASS',   '2014', 'White', '70 138 km', 'FOB 11 771 USD'],
        ['2.jpeg', 'TOYOTA VOXY',               '2014', 'Silver', '79 913 km', 'FOB 12 576 USD'],
        ['3.jpeg', 'SUBARU LEGACY Eyesite G Package',   '2013', 'White', '101 157 km', 'FOB 6 051 USD'],
        ['4.jpeg', 'AUDI A3 Sports Back',   '2013', 'White', '79 311 km', 'FOB 6 437 USD'],
    ]


    sheet.append(['Photo', 'Car Mark',   'Year', 'Color', 'Mileage', 'Price'])
    for cell in ['A', 'B', 'C', 'D', 'E', 'F']:
        sheet[cell + '1'].font = Font(bold=True, size=14)

    counter = 2
    for row in rows: # Добавление строк из картеджа в документ
        # print(row)
        sheet.row_dimensions[counter].height = 90
        cellID = 'A' + str(counter)

        img = openpyxl.drawing.image.Image('images/' + row[0])
        img.height = 120 # insert image height in pixels as float or int (e.g. 305.5)
        img.width = 200 # insert image width in pixels as float or int (e.g. 405.8)
        img.anchor = cellID  # where you want image to be anchored/start from
        sheet.add_image(img)  # adding in the image

        # sheet.cell(row=counter, column=1).value = row[0] # Чтение из ячейки в переменную
        sheet.cell(row=counter, column=2).value = row[1] # Чтение из ячейки в переменную
        sheet.cell(row=counter, column=3).value = row[2] # Чтение из ячейки в переменную
        sheet.cell(row=counter, column=4).value = row[3] # Чтение из ячейки в переменную
        sheet.cell(row=counter, column=5).value = row[4] # Чтение из ячейки в переменную
        sheet.cell(row=counter, column=6).value = row[5] # Чтение из ячейки в переменную

        counter += 1

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    sheet.column_dimensions['A'].width = 30 # Картинка
    # a1 = sheet['A5'] # Чтение из ячейки в переменную
    # a2 = sheet.cell(row=3, column=1) # Чтение из ячейки в переменную
    # print(a1.value)
    # print(a2.value)

    # cells = sheet['A5': 'B7'] # Область ячеек
    # for c1, c2 in cells: # Выдача содержимого ячеек на экран
    #     print("{0:8} {1:8}".format(c1.value, c2.value))
    # print()

    # # Перебор строк нужной области
    # for row in sheet.iter_rows(min_row=5, min_col=1, max_row=10, max_col=3):
    #     for cell in row:
    #         print(cell.value, end=" ")
    #     print()
    # print()
    #
    # # Перебор колонок нужной области
    # for col in sheet.iter_cols(min_row=5, min_col=1, max_row=10, max_col=3):
    #     for cell in col:
    #         print(cell.value, end=" ")
    #     print()

    # cell = sheet.cell(row=11, column=3) # Ячейка
    # cell.value = "=SUM(C5:C10)" # Вставка формулы в ячейку
    # cell.font = Font(name='Calibri', size=20, bold=True) # Установка фонта ячейки
    # sheet.add_image(Image('1.jpeg'), 'B14') # Добавление картинки в ячейку
    # sheet.row_dimensions[3].height = 40 # Установка высота 3 строки
    # sheet.column_dimensions['B'].width = 40 # Установка ширины столбца "B"

    my_file = "sample.xlsx" # Имя файла
    book.save(my_file) # Сохранение файла на диск
    # os.startfile(my_file) # Запуск файла (открытие MS Excel с созданным документом)
    opener = "open" if sys.platform == "darwin" else "xdg-open"
    subprocess.call([opener, my_file])

except Exception as a: # Обработка ошибок
     print("Error!")
     print(a)